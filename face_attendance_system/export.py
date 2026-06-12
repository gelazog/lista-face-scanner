"""
Módulo de exportación a Excel.
Genera archivos .xlsx profesionales con openpyxl.
"""

import os
import subprocess
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

from database import obtener_resumen_asistencia, obtener_asistencia_por_rango

# ─── Paleta de colores ────────────────────────────────────────────────────────
COLOR_ENCABEZADO_FONDO  = "1E1E2E"   # Azul oscuro
COLOR_ENCABEZADO_TEXTO  = "00D4AA"   # Verde acento
COLOR_FILA_PAR          = "2A2A3E"   # Fila alternada oscura
COLOR_FILA_IMPAR        = "252535"   # Fila alternada más oscura
COLOR_FILA_PAR_TEXTO    = "DDEEFF"   # Texto claro fila par
COLOR_FILA_IMPAR_TEXTO  = "CCDDEE"
COLOR_PRESENTE          = "1A6640"   # Verde para "Sí"
COLOR_AUSENTE           = "6B1E1E"   # Rojo para "No"
COLOR_TOTALES_FONDO     = "14142A"
COLOR_TOTALES_TEXTO     = "FFD700"   # Dorado para fila de totales

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "exports")


def _estilo_borde_fino() -> Border:
    lado = Side(style="thin", color="444466")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _celda_encabezado(celda, texto: str):
    """Aplica estilo de encabezado a una celda."""
    celda.value = texto
    celda.font  = Font(name="Segoe UI", bold=True, size=11, color=COLOR_ENCABEZADO_TEXTO)
    celda.fill  = PatternFill("solid", fgColor=COLOR_ENCABEZADO_FONDO)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda.border = _estilo_borde_fino()


def _celda_dato(celda, valor, es_par: bool, color_texto: str = None, negrita: bool = False):
    """Aplica estilo de dato a una celda."""
    celda.value = valor
    fondo = COLOR_FILA_PAR if es_par else COLOR_FILA_IMPAR
    texto = color_texto or (COLOR_FILA_PAR_TEXTO if es_par else COLOR_FILA_IMPAR_TEXTO)
    celda.fill  = PatternFill("solid", fgColor=fondo)
    celda.font  = Font(name="Segoe UI", size=10, color=texto, bold=negrita)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = _estilo_borde_fino()


def exportar_asistencia(
    fecha_inicio: str,
    fecha_fin:    str,
    modo:         str = "resumen"
) -> str:
    """
    Genera un archivo Excel de asistencia y lo devuelve como ruta absoluta.

    Parámetros:
        fecha_inicio: "YYYY-MM-DD"
        fecha_fin:    "YYYY-MM-DD"
        modo: "resumen"  → un registro por usuario (presente/ausente)
              "detallado" → un registro por detección

    Devuelve la ruta del archivo generado.
    """
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H%M")
    nombre_arch = f"Asistencia_{timestamp}.xlsx"
    ruta_arch   = os.path.join(EXPORTS_DIR, nombre_arch)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # ── Fila de título ────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    titulo = ws["A1"]
    titulo.value = f"Registro de Asistencia  |  {fecha_inicio}  →  {fecha_fin}"
    titulo.font  = Font(name="Segoe UI", bold=True, size=14, color=COLOR_ENCABEZADO_TEXTO)
    titulo.fill  = PatternFill("solid", fgColor="0D0D1A")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Encabezados ───────────────────────────────────────────────────────────
    columnas = ["#", "Nombre", "Apellido", "Fecha", "Hora", "Presente"]
    for col_idx, nombre_col in enumerate(columnas, start=1):
        _celda_encabezado(ws.cell(row=2, column=col_idx), nombre_col)
    ws.row_dimensions[2].height = 22

    # ── Datos ─────────────────────────────────────────────────────────────────
    if modo == "resumen":
        registros = obtener_resumen_asistencia(fecha_inicio, fecha_fin)
        filas_datos = []
        for r in registros:
            filas_datos.append({
                "nombre":   r["nombre"],
                "apellido": r["apellido"],
                "fecha":    fecha_inicio if r["presente"] else "—",
                "hora":     "—",
                "presente": "Sí" if r["presente"] else "No"
            })
    else:
        registros_det = obtener_asistencia_por_rango(fecha_inicio, fecha_fin)
        filas_datos = []
        for r in registros_det:
            filas_datos.append({
                "nombre":   r["nombre"],
                "apellido": r["apellido"],
                "fecha":    r["fecha"],
                "hora":     r["hora"],
                "presente": "Sí"
            })

    total_presentes = 0
    total_ausentes  = 0

    for idx, fila in enumerate(filas_datos, start=1):
        fila_excel = idx + 2  # Empieza en fila 3
        es_par     = (idx % 2 == 0)
        es_presente = fila["presente"] == "Sí"

        if es_presente:
            total_presentes += 1
        else:
            total_ausentes += 1

        color_presente = COLOR_PRESENTE if es_presente else COLOR_AUSENTE

        _celda_dato(ws.cell(fila_excel, 1), idx, es_par)
        _celda_dato(ws.cell(fila_excel, 2), fila["nombre"],   es_par)
        _celda_dato(ws.cell(fila_excel, 3), fila["apellido"], es_par)
        _celda_dato(ws.cell(fila_excel, 4), fila["fecha"],    es_par)
        _celda_dato(ws.cell(fila_excel, 5), fila["hora"],     es_par)
        _celda_dato(ws.cell(fila_excel, 6), fila["presente"], es_par,
                    color_texto=color_presente, negrita=True)
        ws.row_dimensions[fila_excel].height = 18

    # ── Fila de totales ───────────────────────────────────────────────────────
    total_registros = len(filas_datos)
    fila_totales    = total_registros + 3

    ws.merge_cells(f"A{fila_totales}:C{fila_totales}")
    celda_etiqueta = ws[f"A{fila_totales}"]
    celda_etiqueta.value = "TOTALES"
    celda_etiqueta.font  = Font(name="Segoe UI", bold=True, size=11, color=COLOR_TOTALES_TEXTO)
    celda_etiqueta.fill  = PatternFill("solid", fgColor=COLOR_TOTALES_FONDO)
    celda_etiqueta.alignment = Alignment(horizontal="center", vertical="center")
    celda_etiqueta.border = _estilo_borde_fino()

    # Columna D: Total registros
    _celda_totales(ws.cell(fila_totales, 4), f"Total: {total_registros}")
    # Columna E: Presentes
    _celda_totales(ws.cell(fila_totales, 5), f"Presentes: {total_presentes}")
    # Columna F: Ausentes / Porcentaje
    porcentaje = (total_presentes / total_registros * 100) if total_registros else 0
    _celda_totales(ws.cell(fila_totales, 6), f"Ausentes: {total_ausentes} | {porcentaje:.1f}% asist.")
    ws.row_dimensions[fila_totales].height = 22

    # ── Anchos de columna ─────────────────────────────────────────────────────
    anchos = {"A": 5, "B": 18, "C": 18, "D": 14, "E": 10, "F": 12}
    for col_letra, ancho in anchos.items():
        ws.column_dimensions[col_letra].width = ancho

    # ── Filtro automático ─────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:F{total_registros + 2}"

    # ── Inmovilizar encabezados ────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Hoja de metadatos ─────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Información")
    _hoja_metadatos(ws_meta, fecha_inicio, fecha_fin, total_registros, total_presentes)

    wb.save(ruta_arch)
    return ruta_arch


def _celda_totales(celda, valor: str):
    """Estilo para celdas de la fila de totales."""
    celda.value = valor
    celda.font  = Font(name="Segoe UI", bold=True, size=10, color=COLOR_TOTALES_TEXTO)
    celda.fill  = PatternFill("solid", fgColor=COLOR_TOTALES_FONDO)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = _estilo_borde_fino()


def _hoja_metadatos(ws, fecha_inicio: str, fecha_fin: str, total: int, presentes: int):
    """Agrega hoja con metadatos de la exportación."""
    ws.sheet_view.showGridLines = False
    datos = [
        ("Generado por",    "Sistema de Asistencia Facial"),
        ("Fecha inicio",    fecha_inicio),
        ("Fecha fin",       fecha_fin),
        ("Total registros", total),
        ("Presentes",       presentes),
        ("Ausentes",        total - presentes),
        ("Exportado el",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    for fila_idx, (clave, valor) in enumerate(datos, start=1):
        ws.cell(fila_idx, 1).value = clave
        ws.cell(fila_idx, 1).font  = Font(bold=True, name="Segoe UI", size=10)
        ws.cell(fila_idx, 2).value = valor
        ws.cell(fila_idx, 2).font  = Font(name="Segoe UI", size=10)


def abrir_carpeta_exports():
    """Abre el explorador de archivos en la carpeta de exports."""
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    if sys.platform == "win32":
        os.startfile(EXPORTS_DIR)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", EXPORTS_DIR])
    else:
        subprocess.Popen(["xdg-open", EXPORTS_DIR])
